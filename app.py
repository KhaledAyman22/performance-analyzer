"""
Ticker Performance Report
Run: streamlit run app.py
"""

import re
from datetime import date, timedelta

import pandas as pd
import streamlit as st
from tvDatafeed import Interval, TvDatafeed

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Ticker Performance Report",
    page_icon="📊",
    layout="wide",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    div[data-testid="stMetricValue"] { font-size: 1.8rem; font-weight: 700; }
    div[data-testid="stMetricLabel"] { font-size: 0.8rem; color: #888; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── TV CONNECTION ─────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner="Connecting to TradingView …")
def _connect() -> TvDatafeed:
    return TvDatafeed()


tv = _connect()


# ── DATA HELPERS ──────────────────────────────────────────────────────────────
def _n_bars(target: date) -> int:
    """Estimate daily bars needed to reach target_date from today."""
    days = max((date.today() - target).days, 0)
    return max(int(days * 5 / 7) + 100, 150)


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_hist(exchange: str, symbol: str, target: date) -> pd.DataFrame | None:
    return tv.get_hist(
        symbol=symbol,
        exchange=exchange,
        interval=Interval.in_daily,
        n_bars=_n_bars(target),
    )


def price_on_or_before(df: pd.DataFrame, target: date) -> tuple[float, date]:
    """Return (close, actual_date) for the last trading day <= target."""
    ts = pd.Timestamp(target)
    candidates = df[df.index <= ts]
    row = candidates.iloc[-1] if not candidates.empty else df.iloc[0]
    return float(row["close"]), row.name.date()


def latest_close(df: pd.DataFrame) -> tuple[float, date]:
    row = df.iloc[-1]
    return float(row["close"]), row.name.date()


# ── QUICK-ADD PARSER ──────────────────────────────────────────────────────────
def parse_list(text: str) -> list[dict]:
    """Parse 'EXCHANGE:TICKER|price' entries (comma- or newline-separated)."""
    out = []
    for part in re.split(r"[,\n]+", text):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^([A-Za-z0-9]+):([A-Za-z0-9._-]+)(?:\|([0-9.]+))?$", part)
        if m:
            out.append(
                {
                    "Exchange": m.group(1).upper(),
                    "Ticker": m.group(2).upper(),
                    "Listing Price": float(m.group(3)) if m.group(3) else None,
                }
            )
    return out


# ── SESSION STATE ─────────────────────────────────────────────────────────────
if "tickers_df" not in st.session_state:
    st.session_state.tickers_df = pd.DataFrame(
        [{"Exchange": "EGX", "Ticker": "COMI", "Listing Price": None}]
    )

# ── HEADER ────────────────────────────────────────────────────────────────────
st.title("📊 Ticker Performance Report")
st.caption(
    "Powered by TradingView · EGX · Tadawul · NYSE · NASDAQ · LSE · DFM · ADX · and more"
)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")

    report_title = st.text_input("Report Title", "Performance Report")

    listing_date = st.date_input(
        "Global Listing Date",
        value=date.today(),
        help=(
            "Used for tickers with no manual listing price — "
            "fetches the closing price on (or the first trading day after) this date."
        ),
    )

    st.divider()
    st.markdown(
        "**Input format**\n\n"
        "`EXCHANGE:TICKER|price`\n\n"
        "Price is optional. Leave blank to auto-fetch the closing price "
        "on the Global Listing Date.\n\n"
        "**Examples**\n"
        "```\n"
        "EGX:COMI|120\n"
        "EGX:HRHO\n"
        "TADAWUL:2222|32.5\n"
        "NYSE:NVDA|210\n"
        "LSE:SHEL\n"
        "DFM:EMAAR\n"
        "```"
    )

# ── QUICK ADD ─────────────────────────────────────────────────────────────────
with st.expander("⚡ Quick Add — paste a list"):
    raw = st.text_area(
        "One per line or comma-separated  ·  format: `EXCHANGE:TICKER|price`",
        placeholder="EGX:COMI|120\nEGX:HRHO|27\nNYSE:NVDA|210\nTADAWUL:2222",
        height=130,
    )
    if st.button("Parse & Add to table"):
        parsed = parse_list(raw)
        if parsed:
            new_rows = pd.DataFrame(parsed)
            st.session_state.tickers_df = pd.concat(
                [st.session_state.tickers_df, new_rows], ignore_index=True
            )
            st.success(f"Added {len(parsed)} ticker(s).")
            st.rerun()
        else:
            st.error(
                "No valid entries found. "
                "Expected format: `EXCHANGE:TICKER` or `EXCHANGE:TICKER|price`"
            )

# ── TICKER TABLE ──────────────────────────────────────────────────────────────
st.subheader("Tickers")

edited_df = st.data_editor(
    st.session_state.tickers_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "Exchange": st.column_config.TextColumn(
            "Exchange",
            width="medium",
            help="TradingView exchange code — EGX, NASDAQ, NYSE, TADAWUL, DFM, ADX, LSE, EURONEXT …",
        ),
        "Ticker": st.column_config.TextColumn(
            "Ticker",
            width="medium",
            help="Symbol exactly as shown on TradingView (e.g. COMI, NVDA, 2222)",
        ),
        "Listing Price": st.column_config.NumberColumn(
            "Listing Price (optional)",
            format="%.4f",
            help="Leave blank → auto-fetch close price on the Global Listing Date",
        ),
    },
    key="ticker_editor",
)
st.session_state.tickers_df = edited_df

col_run, col_clear = st.columns([5, 1])
run_clicked = col_run.button(
    "🚀 Generate Report", type="primary", use_container_width=True
)
if col_clear.button("Clear Results", use_container_width=True):
    st.session_state.pop("results", None)
    st.session_state.pop("errors", None)
    st.rerun()

# ── GENERATE ──────────────────────────────────────────────────────────────────
if run_clicked:
    rows = (
        edited_df.copy()
        .dropna(subset=["Exchange", "Ticker"])
        .query("Exchange.str.strip() != '' and Ticker.str.strip() != ''")
        .to_dict("records")
    )

    if not rows:
        st.warning("Add at least one ticker first.")
    else:
        results, errors = [], []
        progress_bar = st.progress(0.0, text="Starting …")

        for idx, row in enumerate(rows):
            exchange = str(row["Exchange"]).strip().upper()
            symbol = str(row["Ticker"]).strip().upper()
            progress_bar.progress(
                idx / len(rows), text=f"Fetching {exchange}:{symbol} …"
            )

            entry = {"Exchange": exchange, "Ticker": symbol}

            try:
                hist = fetch_hist(exchange, symbol, listing_date)

                if hist is None or hist.empty:
                    entry["Error"] = "Symbol not found on TradingView"
                    errors.append(entry)
                    continue

                # Current (latest) price
                cp, cp_date = latest_close(hist)
                entry["Current Price"] = round(cp, 4)
                entry["As of"] = cp_date

                # Listing price — manual or auto-fetched
                manual_price = row.get("Listing Price")
                if manual_price and not pd.isna(manual_price):
                    lp = float(manual_price)
                    actual_listing_date = listing_date
                    source_label = "Manual"
                else:
                    lp, actual_listing_date = price_on_or_before(hist, listing_date)
                    source_label = f"TV close ({actual_listing_date})"

                entry["Listing Price"] = round(lp, 4)
                entry["Listing Date"] = actual_listing_date
                entry["Price Source"] = source_label

                change = cp - lp
                change_pct = (change / lp) * 100
                entry["Abs Change"] = round(change, 4)
                entry["Change %"] = round(change_pct, 2)
                entry["Status"] = (
                    "🟢 Profit"
                    if change_pct > 0
                    else ("🔴 Loss" if change_pct < 0 else "⚪ Flat")
                )
                entry["Chart"] = (
                    f"https://www.tradingview.com/chart/?symbol={exchange}:{symbol}"
                )

            except Exception as exc:
                entry["Error"] = str(exc)
                errors.append(entry)
                continue

            results.append(entry)

        progress_bar.progress(1.0, text=f"Done — {len(results)} ticker(s) processed.")

        st.session_state.results = results
        st.session_state.errors = errors
        st.session_state.report_title = report_title
        st.session_state.report_date = date.today()

# ── DISPLAY RESULTS ───────────────────────────────────────────────────────────
if st.session_state.get("results"):
    st.divider()

    title_line = st.session_state.get("report_title", "Performance Report")
    gen_date = st.session_state.get("report_date", "")
    st.subheader(f"{title_line}  ·  {gen_date}")

    df = pd.DataFrame(st.session_state["results"])

    # ── Summary metrics ───────────────────────────────────────────────────────
    if "Change %" in df.columns:
        valid = df.dropna(subset=["Change %"])
        n_profit = int((valid["Change %"] > 0).sum())
        n_loss = int((valid["Change %"] < 0).sum())
        avg_pct = valid["Change %"].mean() if len(valid) else 0
        best = valid.loc[valid["Change %"].idxmax()] if len(valid) else None
        worst = valid.loc[valid["Change %"].idxmin()] if len(valid) else None

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Total Tickers", len(df))
        m2.metric("Profitable 🟢", n_profit)
        m3.metric("In Loss 🔴", n_loss)
        m4.metric("Avg Performance", f"{avg_pct:+.2f}%")
        if best is not None:
            m5.metric(
                f"Best · {best['Exchange']}:{best['Ticker']}",
                f"{best['Change %']:+.2f}%",
            )
        if worst is not None:
            m6.metric(
                f"Worst · {worst['Exchange']}:{worst['Ticker']}",
                f"{worst['Change %']:+.2f}%",
            )

    # ── Results table ─────────────────────────────────────────────────────────
    ordered_cols = [
        "Status",
        "Exchange",
        "Ticker",
        "Listing Date",
        "Listing Price",
        "Current Price",
        "As of",
        "Abs Change",
        "Change %",
        "Price Source",
        "Chart",
    ]
    display_cols = [c for c in ordered_cols if c in df.columns]

    st.data_editor(
        df[display_cols],
        disabled=True,
        use_container_width=True,
        height=min(80 + 35 * len(df), 650),
        column_config={
            "Chart": st.column_config.LinkColumn(
                "📈 Chart",
                display_text="Open ↗",
                width="small",
            ),
            "Change %": st.column_config.NumberColumn(
                "Change %",
                format="%.2f%%",
                width="small",
            ),
            "Abs Change": st.column_config.NumberColumn(
                "Abs Change",
                format="%.4f",
                width="small",
            ),
            "Listing Price": st.column_config.NumberColumn(
                "Listing Price",
                format="%.4f",
            ),
            "Current Price": st.column_config.NumberColumn(
                "Current Price",
                format="%.4f",
            ),
            "Status": st.column_config.TextColumn("", width="small"),
        },
    )

    # ── Errors ────────────────────────────────────────────────────────────────
    if st.session_state.get("errors"):
        errs = st.session_state["errors"]
        with st.expander(f"⚠️ {len(errs)} ticker(s) could not be fetched"):
            st.dataframe(pd.DataFrame(errs), use_container_width=True)

    # ── Export ────────────────────────────────────────────────────────────────
    st.divider()
    export_cols = [c for c in ordered_cols if c in df.columns and c != "Chart"]
    export_df = df[export_cols].copy()

    dl1, dl2 = st.columns(2)
    dl1.download_button(
        label="📥 Download CSV",
        data=export_df.to_csv(index=False, encoding="utf-8-sig"),
        file_name=f"performance_report_{gen_date}.csv",
        mime="text/csv",
        use_container_width=True,
    )

    # Excel export (requires openpyxl)
    try:
        import io
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Performance")
            ws = writer.sheets["Performance"]

            # Header style
            header_fill = PatternFill("solid", fgColor="163D24")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="C9A84C")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Color-code Change % column
            chg_col_idx = export_cols.index("Change %") + 1 if "Change %" in export_cols else None
            if chg_col_idx:
                for row_cells in ws.iter_rows(
                    min_row=2, max_row=ws.max_row,
                    min_col=chg_col_idx, max_col=chg_col_idx
                ):
                    for cell in row_cells:
                        if cell.value is not None:
                            color = "27AE60" if cell.value > 0 else "E74C3C"
                            cell.font = Font(bold=True, color=color)

            # Auto-fit columns
            for i, col in enumerate(export_cols, 1):
                max_len = max(len(str(col)), 10)
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 30)

        dl2.download_button(
            label="📊 Download Excel",
            data=buf.getvalue(),
            file_name=f"performance_report_{gen_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except ImportError:
        dl2.info("Install `openpyxl` for Excel export.")
